import os
import glob
from run_anp_matching import match_anp_with_image
from run_clip_iqa import run_clip_iqa
from nima.evaluate_inception_resnet import extract_features
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
import os
import sys


def load_anp_emotion_dict(txt_path):
    """加载 ANP:emotion score 字典的嵌套结构"""
    emotion_dict = {}
    with open(txt_path, "r", encoding="utf-8") as f:
        current_anp = None
        for line in f:
            line = line.strip()
            if line.startswith("ANP:"):
                current_anp = line.split("ANP:")[1].strip()
                emotion_dict[current_anp] = {}
            elif current_anp and ":" in line:
                try:
                    emotion, value = line.split(":")
                    emotion = emotion.strip()
                    value = float(value.strip())
                    emotion_dict[current_anp][emotion] = value
                except ValueError:
                    continue
    return emotion_dict


def get_max_emotion_score_for_anp(anp_name, emotion_dict):
    """返回某个ANP在情绪分数中得分最高的 (情绪, 分数)"""
    anp_key = anp_name.replace(" ", "_")
    if anp_key in emotion_dict:
        emotion_scores = emotion_dict[anp_key]
        if emotion_scores:
            return max(emotion_scores.items(), key=lambda x: x[1])
    return ("N/A", None)


def save_folder_results_to_excel(summary_excel_path, subdir_name, results):
    """
    计算并保存每个文件夹中以 *-1.png 和 *-2.png 结尾图片的平均指标值到 Excel。
    不记录图片数量。
    """
    def compute_avg(filtered_results):
        if not filtered_results:
            return (0, 0, 0, 0)
        return (
            round(sum(r["emotion_score"] for r in filtered_results) / len(filtered_results), 4),
            round(sum(r["iqa_avg"] for r in filtered_results) / len(filtered_results), 4),
            round(sum(r["nima_mean"] for r in filtered_results) / len(filtered_results), 4),
            round(sum(r["nima_std"] for r in filtered_results) / len(filtered_results), 4),
        )

    group_1 = [r for r in results if r["image"].endswith("-1.png")]
    group_2 = [r for r in results if r["image"].endswith("-2.png")]

    avg_1 = compute_avg(group_1)
    avg_2 = compute_avg(group_2)

    # 创建或加载 Excel
    if os.path.exists(summary_excel_path):
        try:
            wb = load_workbook(summary_excel_path)
            sheet = wb.active
        except InvalidFileException:
            wb = Workbook()
            sheet = wb.active
            sheet.append([
                "帖子名称",
                "Emotion Avg (原图)", "IQA Avg (原图)", "NIMA Mean (原图)", "NIMA Std (原图)",
                "Emotion Avg (漫画)", "IQA Avg (漫画)", "NIMA Mean (漫画)", "NIMA Std (漫画)"
            ])
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append([
            "帖子名称",
            "Emotion Avg (原图)", "IQA Avg (原图)", "NIMA Mean (原图)", "NIMA Std (原图)",
            "Emotion Avg (漫画)", "IQA Avg (漫画)", "NIMA Mean (漫画)", "NIMA Std (漫画)"
        ])

    # 添加一行数据
    sheet.append([
        subdir_name,
        *avg_1,
        *avg_2
    ])

    wb.save(summary_excel_path)


def get_all_max_emotions_for_top_anps(anp_results, emotion_dict):
    """处理每个返回的ANP，获取其情绪得分最高项"""
    results = []
    for anp, anp_score in anp_results[:10]:
        emotion, score = get_max_emotion_score_for_anp(anp, emotion_dict)
        results.append((anp, anp_score, emotion, score))
    return results

def wtd_avg(max_emotions):
    weighted_sum = 0.0
    weight_total = 0.0
    for anp, anp_score, _, emotion_score in max_emotions:
        if emotion_score is not None:
            weighted_sum += anp_score * emotion_score
            weight_total += anp_score

    if weight_total > 0:
        weighted_emotion_score = weighted_sum / weight_total
    else:
        weighted_emotion_score = 0  # 或者设为 0.0 或 "N/A"

    return weighted_emotion_score


def save_txt_log(results, txt_path, subdir_name):
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(f"帖子名称：{subdir_name}\n")
        f.write(f"图片数量：{len(results)}\n")
        f.write(f"{'Image':20s} | {'Emotion Score':14s} | {'IQA Avg':10s} | {'NIMA Mean':10s} | {'NIMA Std':10s}\n")
        f.write("-" * 75 + "\n")
        for entry in results:
            f.write(
                f"{entry['image']:20s} | "
                f"{entry['emotion_score'] if entry['emotion_score'] is not None else '':14.4f} | "
                f"{entry['iqa_avg'] if entry['iqa_avg'] is not None else '':10.4f} | "
                f"{entry['nima_mean']:10.4f} | "
                f"{entry['nima_std']:10.4f}\n"
            )


save_path = "result.xlsx"
emotion_scores_dict = load_anp_emotion_dict("anp_emotion_scores.txt")

img_dir = os.path.join(os.getcwd(), "imgs")
dir_list = [f for f in os.listdir(img_dir) if os.path.isdir(os.path.join(img_dir, f))]


for sub_dir in dir_list:
    # print(sub_dir)
    # exit()
    full_dir = os.path.join(img_dir, sub_dir)
    jpg_files = glob.glob(os.path.join(full_dir, "*.png"))
    jpg_files.sort(key=lambda x: os.path.basename(x))
    
    print(f"Processing folder: {sub_dir}")
    results = []
    forlder_results = []

    for jpg in jpg_files:
        print(f"NOW PROCESSING {sub_dir} / {[os.path.basename(jpg)]}")

        # ANP
        anp_results = match_anp_with_image(jpg)
        max_emotions = get_all_max_emotions_for_top_anps(anp_results, emotion_scores_dict)
        wtd_avg_score = wtd_avg(max_emotions)

        # print("Top 10 ANPs and their most intense emotions:")
        # for anp, anp_score, emotion, emotion_score in max_emotions:
        #     print(f"{anp:30s}: original score: {anp_score:.4f} → Emotion Type: {emotion:10s}: {emotion_score:.4f}")
        # print(wtd_avg_score)
        # print()

        # IQA
        iqa_score = run_clip_iqa(jpg)
        iqa_score = {k: float(v) for k, v in iqa_score.items()}  # 转换 tensor → float

        iqa_values= list(iqa_score.values())
        if iqa_values:
            iqa_avg = sum(iqa_values) / len(iqa_values)
        else:
            iqa_avg = -1

        # print("CLIP-IQA score:")
        # for k, v in iqa_score.items():
        #     print(f"{k}: {v}")
        # print()

        # NIMA
        nima_mean, nima_std = extract_features(jpg, resize=True)
        # print("NIMA Score : %0.3f +- (%0.3f)" % (nima_mean, nima_std))
        # print()

        results.append({
            "image": os.path.basename(jpg),
            "emotion_score": wtd_avg_score,
            "iqa_avg": iqa_avg,
            "nima_mean": nima_mean,
            "nima_std": nima_std
        })

        txt_path = os.path.join(full_dir, f"{sub_dir}_结果日志.txt")
        save_txt_log(results, txt_path, sub_dir)

    # 保存到总表格中（主目录下）
    # print(results)
    # exit()
    summary_excel = os.path.join(os.getcwd(), "result.xlsx")
    save_folder_results_to_excel(summary_excel, sub_dir, results)
    print("=" * 60) 



